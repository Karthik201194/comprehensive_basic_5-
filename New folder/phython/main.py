from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

# Function to read data from Excel sheet
def read_excel_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

# Launch WebDriver and open URL
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.saucedemo.com/")
expected_title = "Swag Labs"

# Verify title of the page
actual_title = driver.title
if actual_title == expected_title:
    print("Title verified:", actual_title)
else:
    print("Title verification failed:", actual_title)

test_data = read_excel_data("testdata.xlsx")


for username, password, expected_result in test_data:
    
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()

    
    if expected_result == "success":
        if driver.title == expected_title:
            print("Login successful for username:", username)
        else:
            print("Login failed for username:", username, "- Unexpected page title:", driver.title)
    else:
    
        error_message = driver.find_element(By.CSS_SELECTOR, "h3[data-test='error']").text
        if error_message:
            print("Login failed for username:", username, "- Error message:", error_message)
        else:
            print("Login failed for username:", username, "- No error message displayed")


    driver.find_element(By.ID, "user-name").clear()
    driver.find_element(By.ID, "password").clear()

driver.quit()
